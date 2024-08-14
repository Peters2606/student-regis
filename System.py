from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "#06283D"
framebg="#EDEDED"
framefg="#06283D"

root=Tk()
root.title("Student Registration System") #title of the window
root.geometry("1250x700+210+100")
root.config(bg=background)

file=pathlib.Path("student_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="Religion"
    sheet['H1']="skill"
    sheet['I1']="Father name"
    sheet['J1']="Mother name"
    sheet['K1']="Father occupation"
    sheet['L1']="Mother occupation"

    file.save("student_data.xlsx")

#exit window
def exit():
    root.destroy()

#showimage
def showimage():
    global filename
    global img
    filename=filedialog.askopenfile(initialdir=os.getcwd(),title="select image file",filetypes=(("JPG File","*.jpg"),("PNG File","*.png"),("All Files","*.txt")))

    img = (Image.open(filename.name))
    resized_image = img.resize((200, 200))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#registration number  automatic
def regno():
    file=openpyxl.load_workbook("student_data.xlsx")
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try :
        registration.set(max_row_value+1)

    except :
        registration.set(1)

#clear
def clear():
    global img

    Name.set("")
    Class.set("Select class")
    DOB.set("")
    registration.set("")
    Religion.set("")
    Skill.set("")
    F_name.set("")
    M_name.set("")
    Father_Occupation.set("")
    Mother_Occupation.set("")

    regno()

    saveButton.config(state="normal")

    img1=PhotoImage(file="images/upload2.png")
    lbl.config(image=img1)
    lbl.image=img1

    img=""

#save
def save():
    R1=registration.get()
    N1=Name.get()
    C1=Class.get()
    try :
        G1=gender
    except:
        messagebox.showerror("Error","Please select gender")
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skill.get()
    fathername=F_name.get()
    mothername=M_name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()

    if N1=="" or C1=="" or G1=="" or D2=="" or D1=="" or Re1=="" or S1=="" or fathername=="" or mothername=="" or F1=="" or M1=="":
        messagebox.showerror("Error","Few data is missing")
    else :
        file=openpyxl.load_workbook("student_data.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=Re1)
        sheet.cell(column=8,row=sheet.max_row,value=S1)
        sheet.cell(column=9,row=sheet.max_row,value=fathername)
        sheet.cell(column=10,row=sheet.max_row,value=mothername)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)
        
        file.save("student_data.xlsx")
        messagebox.showinfo("Success","Data saved successfully")
        try :
            img.save("Student Images/"+str(R1)+".jpg")
        except :
            messagebox.showerror("Error","Image not saved")

        messagebox.showinfo("Success","Data saved successfully")

        clear() #clear box and image section

        regno() # it will recheck the registration number

    print(R1,N1,C1,G1,D2,D1,Re1,S1,fathername,mothername,F1,M1)

#search
def search():
    text=Search.get() #mengambil data dari entry box
    clear() # menghilangkan semua data di dalam entry box dan lainnya
    saveButton.config(state="disable") #ketika btn search di  klik, btn save di disable

    file=openpyxl.load_workbook("student_data.xlsx")
    sheet=file.active
    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    try :
        print(str(name))
    except :
        messagebox.showerror("Error","No data found")
        #reg_no_position menunjukkan seperti A1,A2,...,An
        #reg_number hanya menunjukkan angka setlah A1,A2,...,An seperti 1,2,3,...,n
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

    registration.set(x1)
    Name.set(x2)
    Class.set(x3)
    if x4=="Male":
        R2.select()
    else :
        R1.select()
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    F_name.set(x9)
    M_name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)

    img = (Image.open("Student Images/"+str(x1)+".jpg")) #mengambil gambar dengan nama yang sama dengan registration number
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#update
def Update():
    R1=registration.get()
    N1=Name.get()
    C1=Class.get()
    selection()
    G1=gender
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skill.get()
    fathername=F_name.get()
    mothername=M_name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()

    file=openpyxl.load_workbook("student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            print(reg_number)

    #sheet.cell(column=1,row=int(reg_number),value=R1) //now,noone can update registration number,it will remain same
    sheet.cell(column=2,row=int(reg_number),value=N1) 
    sheet.cell(column=3,row=int(reg_number),value=C1) 
    sheet.cell(column=4,row=int(reg_number),value=G1) 
    sheet.cell(column=5,row=int(reg_number),value=D2) 
    sheet.cell(column=6,row=int(reg_number),value=D1) 
    sheet.cell(column=7,row=int(reg_number),value=Re1) 
    sheet.cell(column=8,row=int(reg_number),value=S1) 
    sheet.cell(column=9,row=int(reg_number),value=fathername) 
    sheet.cell(column=10,row=int(reg_number),value=mothername) 
    sheet.cell(column=11,row=int(reg_number),value=F1) 
    sheet.cell(column=12,row=int(reg_number),value=M1) 

    file.save(r'student_data.xlsx')

    try :
        img.save("Student Images/"+str(R1)+".jpg")
    except :
        pass

    messagebox.showinfo("Update","Updated Successfully")

    clear()

    
    


#gender
def selection():
    global gender
    value=Radio.get()
    if value==1:
        gender="Male"
    else :
        gender="Female"

#top Frames
Label(root,text="Email: petersgabriel704.gmail.com",width=10,height=3,bg="#f0687c",anchor="e").pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#c36464",fg="white",font="arial 20 bold").pack(side=TOP,fill=X)

#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=15,font="arial 20",bd=2).place(x=850,y=70)
imageicon3=PhotoImage(file="images/search.png")
srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,bg="#68ddfa",font="arial 13 bold",command=search)
srch.place(x=1100,y=73)

imageicon4=PhotoImage(file="images/update.png")
update_button=Button(root,text="Update",compound=LEFT,image=imageicon4,width=123,bg="#68ddfa",font="arial 13 bold",command=Update)
update_button=update_button.place(x=110,y=74)

#registration and date
Label(root,text="Registration No : ",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

registration=IntVar()
Date=StringVar()

reg_entry=Entry(root,textvariable=registration,width=15,font="arial 10").place(x=158,y=150)

regno() #call fungsion regno to generate registration number

today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10").place(x=580,y=150)
Date.set(d1)

#student details
obj = LabelFrame(root, text="Student's Details", font=20, bg=framebg, fg=framefg, bd=2,width=900,height=250, relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name     : ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth : ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender         : ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Class          : ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion      : ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skils           : ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name=StringVar()
name_entry=Entry(obj, textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DOB=StringVar()
dob_entry=Entry(obj, textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)

Radio=IntVar()
R1=Radiobutton(obj,text="Male",bg=framebg,fg=framefg,variable=Radio,value=1,command=selection)
R1.place(x=160,y=150)
R2=Radiobutton(obj,text="Female",bg=framebg,fg=framefg,variable=Radio,value=2,command=selection)
R2.place(x=230,y=150)

Religion=StringVar()
religion_entry=Entry(obj, textvariable=Religion,width=20,font="arial 10")
religion_entry.place(x=630,y=100)

Skill=StringVar()
skill_entry=Entry(obj, textvariable=Skill,width=20,font="arial 10")
skill_entry.place(x=630,y=150)

Class = Combobox(obj,values=["1","2","3","4","5","6","7","8","9","10","11","12"],font="Roboto 10",state="r",width=17)
Class.place(x=630,y=50)
Class.set("Select Class")

#Parent details
obj1 = LabelFrame(root, text="Parent's Details", font=20, bg=framebg, fg=framefg, bd=2,width=900,height=200, relief=GROOVE)
obj1.place(x=30,y=470)

#father
Label(obj1,text="Father's Name  : ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj1,text="Occupation      : ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

F_name=StringVar()
f_name_entry=Entry(obj1, textvariable=F_name,width=20,font="arial 10")
f_name_entry.place(x=160,y=50)

Father_Occupation=StringVar()
FO_entry=Entry(obj1, textvariable=Father_Occupation,width=20,font="arial 10")
FO_entry.place(x=160,y=100)


#mother
Label(obj1,text="Mother's Name  : ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj1,text="Occupation       : ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

M_name=StringVar()
m_name_entry=Entry(obj1, textvariable=M_name,width=20,font="arial 10")
m_name_entry.place(x=630,y=50)

Mother_Occupation=StringVar()
MO_entry=Entry(obj1, textvariable=Mother_Occupation,width=20,font="arial 10")
MO_entry.place(x=630,y=100)

#image
f=Frame(root,bd=3,bg="Black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="images/upload2.png")
lbl=Label(f,bg="Black",image=img)
lbl.place(x=0,y=0)

#buttons
Button(root,text="Upload",width=19,height=2,bg="Lightblue",font="arial 12 bold",command=showimage).place(x=1000,y=370)

saveButton=Button(root,text="Save",width=19,height=2,bg="Lightgreen",font="arial 12 bold",command=save)
saveButton.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,bg="Lightpink",font="arial 12 bold",command=clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,bg="Grey",font="arial 12 bold",command=exit).place(x=1000,y=610)

root.mainloop()
